import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl

from app.charges.models.charge import ChargeStatus, ChargeType


def _seed_charges(create_client_with_business, charge_factory):
    client, business = create_client_with_business(
        full_name="Export Aging Client", id_number="AGING-EXP-1"
    )

    issued_at = date.today() - timedelta(days=40)
    charge_factory(
        client_record_id=client.id,
        business_id=business.id,
        amount=Decimal("250.00"),
        charge_type=ChargeType.CONSULTATION_FEE,
        status=ChargeStatus.ISSUED,
        issued_at=issued_at,
        commit=True,
    )
    return client


def test_aging_excel_exporter(
    client, test_db, advisor_headers, create_client_with_business, charge_factory
):
    crm_client = _seed_charges(create_client_with_business, charge_factory)

    resp = client.get("/api/v1/reports/aging/export?format=excel", headers=advisor_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Title row merged A1:H1
    assert ws["A1"].value is not None
    # Data row for client
    client_names = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
    assert crm_client.full_name in client_names


def test_aging_pdf_exporter(
    client, test_db, advisor_headers, create_client_with_business, charge_factory
):
    _seed_charges(create_client_with_business, charge_factory)

    resp = client.get("/api/v1/reports/aging/export?format=pdf", headers=advisor_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    assert len(resp.content) > 0
