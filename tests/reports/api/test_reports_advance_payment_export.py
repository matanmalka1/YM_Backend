import io
from decimal import Decimal

import openpyxl

from app.advance_payments.services.advance_payment_service import AdvancePaymentService
from app.common.enums import AdvancePaymentFrequency


def _seed_payment(db, client_factory):
    client = client_factory(
        advance_payment_frequency=AdvancePaymentFrequency.MONTHLY,
        advance_rate=Decimal("10"),
    )
    AdvancePaymentService(db).create_payment_for_client(
        client_record_id=client.id,
        period="2026-02",
        period_months_count=1,
        turnover_amount=Decimal("40000"),
        paid_amount=Decimal("1000"),
    )
    db.commit()
    return client


def test_advance_payment_report_excel_export(client, test_db, advisor_headers, client_factory):
    crm_client = _seed_payment(test_db, client_factory)

    resp = client.get(
        "/api/v1/reports/advance-payments/export?format=excel&year=2026", headers=advisor_headers
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[3]]
    assert headers[0] == "מס׳ לקוח"
    assert headers[3] == "צפוי"

    names = [row[1].value for row in ws.iter_rows(min_row=4)]
    assert crm_client.full_name in names


def test_advance_payment_report_pdf_export(client, test_db, advisor_headers, client_factory):
    _seed_payment(test_db, client_factory)

    resp = client.get(
        "/api/v1/reports/advance-payments/export?format=pdf&year=2026", headers=advisor_headers
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    assert resp.content.startswith(b"%PDF")


def test_advance_payment_report_export_rejects_bad_params(client, test_db, advisor_headers):
    resp = client.get(
        "/api/v1/reports/advance-payments/export?format=excel&year=2026&month=13",
        headers=advisor_headers,
    )
    assert resp.status_code == 422

    resp = client.get(
        "/api/v1/reports/advance-payments/export?format=csv&year=2026", headers=advisor_headers
    )
    assert resp.status_code == 422
